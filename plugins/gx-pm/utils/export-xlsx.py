#!/usr/bin/env python3
"""
gx-pm 산출물 마크다운 표 → xlsx 변환 유틸리티

Usage:
    python export-xlsx.py <input.md> [--output output.xlsx]
    python export-xlsx.py --dir <폴더> [--output output.xlsx]

Examples:
    python export-xlsx.py ACT-요구사항정의서.md
    python export-xlsx.py --dir ../결과물/ --output ACT-산출물.xlsx
    python export-xlsx.py ACT-요구사항정의서.md ACT-화면목록표.md --output merged.xlsx
"""

import sys
import re
import os
import io
import argparse
from pathlib import Path

# 목록 표지는 마커 뒤에 공백이 온다("- 항목", "1. 항목").
# 공백을 요구해야 볼드 강조("**부적합 목록**")를 목록으로 오인하지 않는다.
_BULLET_PREFIX = re.compile(r"^([-*+]|\d+\.)\s")

# Windows cp949 인코딩 문제 방지
if sys.stdout.encoding and sys.stdout.encoding.lower().startswith("cp"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


# ──────────────────────────────────────
# 의존성 자동 설치
# ──────────────────────────────────────

def ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        import subprocess
        print("[gx-pm] openpyxl 설치 중...", file=sys.stderr)
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
            stdout=subprocess.DEVNULL,
        )
        import openpyxl
        return openpyxl


# ──────────────────────────────────────
# 산출물별 컬럼 매핑 (공공 양식 순서)
# ──────────────────────────────────────

DOCUMENT_PROFILES = {
    "요구사항정의서": {
        "sheet_name": "AN-02 요구사항정의서",
        "columns": [[
            "제안요청ID", "요구사항ID", "구분", "근거",
            "대분류", "중분류", "소분류",
            "요구내역", "요구사항 상세내역", "수용여부",
        ]],
    },
    "화면목록표": {
        "sheet_name": "DE-03 화면목록표",
        "columns": [[
            "기능구분ID", "기능구분명", "기능ID", "기능명", "SN",
            "화면ID", "화면명", "망구분", "담당자",
            "개발예정일_UI", "개발예정일_LOGIC",
            "개발진행_UI", "개발진행_LOGIC", "신규/변경", "비고",
        ]],
    },
    "프로그램정의서": {
        "sheet_name": "DE-05 프로그램정의서",
        "columns": [[
            "NO", "업무명", "대분류", "중분류", "소분류",
            "ID", "프로그램명", "메뉴_URL", "Package",
            "Controller", "VO", "DAO", "Service", "Impl", "Sql", "JSP",
            "요구사항ID", "화면설계서ID",
        ]],
    },
    "테이블정의서": {
        "sheet_name": "DE-08 테이블정의서",
        "columns": [[
            "테이블명", "엔터티명", "컬럼명", "속성명",
            "데이터타입", "길이", "소수점", "기본값",
            "PK", "FK", "NotNull", "표준 출처",
        ]],
    },
    "인터페이스정의서": {
        "sheet_name": "DE-04 인터페이스정의서",
        "columns": [[
            "인터페이스ID", "인터페이스명", "연동대상 시스템", "연동방식",
            "송수신방향", "주기", "관련 요구사항ID", "비고",
        ]],
    },
    "총괄테스트계획서": {
        "sheet_name": "TE-01 총괄테스트계획서",
        "sheet_names": ["TE-01 테스트 레벨", "TE-01 종료기준", "TE-01 추진체제"],
        "columns": [
            ["레벨", "산출물 코드", "대상", "수행 주체", "검증 대상 요구사항"],
            ["기준", "목표치", "비고"],
            ["역할", "담당", "책임"],
        ],
    },
    "단위테스트계획서": {
        "sheet_name": "DE-13 단위테스트계획서",
        "sheet_names": ["DE-13 단위테스트계획", "DE-13 테스트케이스"],
        "columns": [
            # 시트 1 — 검사기준 체크리스트 (뒤의 27개 검사항목 컬럼은 가변)
            ["기능구분ID", "기능구분명", "기능ID", "기능명",
             "화면ID", "화면명", "사용자구분", "단위테스트ID"],
            # 시트 2 — 테스트케이스
            ["화면ID", "화면명", "단위테스트ID", "테스트케이스ID", "요구사항ID",
             "검사기준 항목", "설계기법", "구분", "사전조건", "입력 데이터",
             "수행 절차", "기대 결과", "우선순위",
             "실행결과", "점검제외 사유", "결함ID", "수행자", "수행일"],
        ],
    },
    "단위테스트결과서": {
        "sheet_name": "IM-03 단위테스트결과서",
        "columns": [[
            "화면ID", "화면명", "구분", "항목번호", "검사항목",
            "계획여부", "검사결과", "부적합내역", "결함ID", "점검제외 사유",
            "테스트일자", "테스트실시자",
        ]],
    },
    "통합테스트시나리오": {
        "sheet_name": "DE-14 통합테스트시나리오",
        "columns": [[
            "통합테스트 ID", "요구사항 ID", "시나리오명", "구분", "사전조건",
            "Step", "사용자구분",
            "예상 입력 (사용자 행위)", "예상 출력 (시스템 결과)", "테스트 데이터",
        ]],
    },
    "통합테스트결과서": {
        "sheet_name": "TE-02 통합테스트결과서",
        "columns": [[
            "시나리오ID", "시나리오명", "구분", "사전조건", "Step", "테스트데이터",
            "판정결과", "부적합내역", "결함ID", "점검제외 사유",
            "테스트일시", "수행자",
        ]],
    },
    "인수테스트결과서": {
        "sheet_name": "TE-06 인수테스트결과서",
        "columns": [[
            "시나리오ID", "시나리오명", "구분", "사전조건", "Step", "테스트데이터",
            "판정결과", "부적합내역", "결함ID", "점검제외 사유",
            "테스트일시", "수행자",
        ]],
    },
    "시스템테스트계획서": {
        "sheet_name": "ST-01 시스템테스트계획서",
        "columns": [[
            "시스템테스트ID", "요구사항ID", "비기능유형", "테스트 항목",
            "측정 방법", "목표치", "테스트 조건", "비고",
        ]],
    },
    "시스템테스트결과서": {
        "sheet_name": "ST-02 시스템테스트결과서",
        "columns": [[
            "시스템테스트ID", "요구사항ID", "비기능유형", "테스트 항목",
            "측정 방법", "목표치", "테스트 조건", "측정값",
            "판정결과", "점검제외 사유", "결함ID",
            "측정일시", "수행자", "증적", "비고",
        ]],
    },
    "결함관리대장": {
        "sheet_name": "DF-01 결함관리대장",
        "columns": [[
            "결함ID", "발견단계", "발견 테스트ID", "화면ID", "요구사항ID",
            "결함 제목", "결함 내역", "재현 절차", "심각도", "우선순위",
            "발견일", "발견자", "상태", "담당자",
            "원인", "조치내역", "조치일", "조치자",
            "재테스트 결과", "재테스트일", "재테스트자", "비고",
        ]],
    },
    "추적매트릭스": {
        "sheet_name": "AN-05 추적매트릭스",
        "columns": [[
            "제안요청ID", "제안요청 내역",
            "요구사항ID", "요구사항 내역", "구분", "수용여부",
            "인터페이스ID", "인터페이스명", "테이블ID",
            "화면ID", "화면명", "프로그램ID", "프로그램명",
            "단위테스트ID", "테스트케이스 수",
            "통합테스트ID", "통합테스트시나리오명",
            "시스템테스트ID", "시스템테스트 항목",
            "잔존 결함", "과업완료여부", "검사기준",
        ]],
    },
}


# ──────────────────────────────────────
# 마크다운 파싱
# ──────────────────────────────────────

def detect_document_type(filename: str) -> str | None:
    """파일명에서 산출물 유형을 감지"""
    for key in DOCUMENT_PROFILES:
        if key in filename:
            return key
    return None


def parse_markdown_tables(text: str) -> list[tuple[str, list[str]]]:
    """마크다운 텍스트에서 (제목, 표 라인 리스트) 쌍을 추출"""
    tables = []
    lines = text.split("\n")
    current_table: list[str] = []
    table_title = ""

    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            if not current_table:
                # 표 직전의 제목(### 등)을 찾는다.
                # 코드펜스(``` / ~~~) 블록은 여는/닫는 펜스 사이를 통째로
                # 건너뛴다 — 탐색 창(budget)도 소모하지 않는다. 코드 블록이
                # 길어도 그 위의 상위 헤딩까지 거슬러 올라갈 수 있어야 한다.
                j = i - 1
                budget = 30
                while j >= 0 and budget > 0:
                    candidate = lines[j].strip()
                    if candidate.startswith(("```", "~~~")):
                        fence = candidate[:3]
                        j -= 1
                        while j >= 0 and not lines[j].strip().startswith(fence):
                            j -= 1
                        j -= 1
                        continue
                    budget -= 1
                    if candidate.startswith("#"):
                        table_title = candidate.lstrip("#").strip()
                        break
                    # 인용문·표·목록은 제목이 아니다 — 계속 거슬러 올라간다.
                    # 목록 마커는 뒤에 공백이 온다("- 항목"). 공백이 없으면
                    # 볼드 강조("**부적합 목록**")이므로 라벨로 인정한다.
                    if candidate.startswith((">", "|")) or _BULLET_PREFIX.match(candidate):
                        j -= 1
                        continue
                    # 산문 한 줄이 시트명이 되는 것을 막는다.
                    # 짧고 문장으로 끝나지 않는 라인만 라벨로 인정한다.
                    if candidate and len(candidate) <= 30 and not candidate.endswith((".", "다", "요", "!", "?")):
                        table_title = candidate
                        break
                    j -= 1
            current_table.append(stripped)
        else:
            if current_table:
                tables.append((table_title, current_table))
                current_table = []
                table_title = ""

    if current_table:
        tables.append((table_title, current_table))

    return tables


def table_lines_to_rows(table_lines: list[str]) -> list[list[str]]:
    """마크다운 표 라인 → 2D 배열 (구분선 제거)"""
    rows = []
    for line in table_lines:
        # 구분선 ( |---|---| ) 건너뛰기
        cells = [c.strip() for c in line.split("|")[1:-1]]
        if cells and not all(re.match(r"^[\s\-:]+$", c) for c in cells):
            rows.append(cells)
    return rows


def is_metadata_table(rows: list[list[str]]) -> bool:
    """메타 정보 표(항목/내용 2열, 판정기준 등)인지 판별 — 산출물 데이터 표와 분리"""
    if not rows:
        return True
    header = [h.strip() for h in rows[0]]
    # 2열짜리 "항목 | 내용" 스타일의 문서 헤더 메타 표
    if len(header) == 2 and header[0] in ("항목", "항목명"):
        return True
    # 2열짜리 참조 표 (판정 | 의미, 키워드 | 값 등)
    if len(header) == 2 and header[0] in ("판정", "키워드", "유형"):
        return True
    # 데이터 행이 2개 이하 (헤더 포함 총 3행 이하)이면 의미 없는 소형 표
    if len(rows) <= 2 and len(header) <= 3:
        return True
    return False


# ──────────────────────────────────────
# xlsx 생성
# ──────────────────────────────────────

def _header_key(rows: list[list[str]]) -> str:
    """표의 헤더를 정규화한 키 — 동일 구조 표 병합용"""
    if not rows:
        return ""
    return "|".join(h.strip() for h in rows[0])


def _matched_set_index(rows: list[list[str]], doc_type: str | None) -> int | None:
    """이 표가 산출물 본문 표라면 매칭된 컬럼 세트의 인덱스를, 아니면 None 을 반환한다.

    프로필 컬럼 세트와 절반 이상 일치하면 본문 표다.
    근거·통계 같은 보조 표는 산출물 시트명을 물려받지 않고 자기 제목을 쓴다.
    """
    if not doc_type or doc_type not in DOCUMENT_PROFILES or not rows:
        return None
    header = {h.strip() for h in rows[0]}
    best: int | None = None
    best_ratio = 0.0
    for index, target in enumerate(DOCUMENT_PROFILES[doc_type]["columns"]):
        if not target:
            continue
        ratio = len([c for c in target if c in header]) / len(target)
        if ratio >= 0.5 and ratio > best_ratio:
            best, best_ratio = index, ratio
    return best


def _reorder_columns(
    rows: list[list[str]], doc_type: str | None
) -> list[list[str]]:
    """DOCUMENT_PROFILES에 정의된 공공 양식 컬럼 순서로 재배열한다.

    프로필에 정의된 컬럼이 마크다운 헤더에 존재하면 해당 순서로 재배열하고,
    프로필에 없는 추가 컬럼은 뒤에 붙인다.
    매칭되는 컬럼이 절반 미만이면 재배열하지 않고 원본을 반환한다.
    """
    if not doc_type or doc_type not in DOCUMENT_PROFILES or not rows:
        return rows

    column_sets = DOCUMENT_PROFILES[doc_type]["columns"]
    header = [h.strip() for h in rows[0]]

    # 마크다운 헤더 → 인덱스 매핑
    header_index: dict[str, int] = {}
    for idx, col in enumerate(header):
        header_index[col] = idx

    # 산출물이 여러 시트 구조를 가질 수 있으므로 가장 잘 맞는 컬럼 세트를 고른다
    best_indices: list[int] = []
    best_ratio = 0.0
    for target_columns in column_sets:
        if not target_columns:
            continue
        indices = [header_index[c] for c in target_columns if c in header_index]
        ratio = len(indices) / len(target_columns)
        if ratio > best_ratio:
            best_ratio, best_indices = ratio, indices

    # 매칭률이 절반 미만이면 재배열 의미 없음 — 원본 반환
    if best_ratio < 0.5:
        return rows

    ordered_indices = list(best_indices)

    # 프로필에 없는 추가 컬럼을 뒤에 붙인다
    used = set(ordered_indices)
    for idx in range(len(header)):
        if idx not in used:
            ordered_indices.append(idx)

    # 모든 행에 재배열 적용
    reordered: list[list[str]] = []
    for row in rows:
        new_row = []
        for idx in ordered_indices:
            new_row.append(row[idx] if idx < len(row) else "")
        reordered.append(new_row)

    return reordered


def create_xlsx(
    file_tables: list[tuple[str, list[tuple[str, list[str]]]]],
    output_path: str,
) -> str:
    """
    file_tables: [(파일명, [(표제목, 표라인[]), ...]), ...]

    동일 파일 내 같은 컬럼 구조의 표는 하나의 시트로 병합한다.
    """
    openpyxl = ensure_openpyxl()
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 스타일 정의
    header_font = Font(name="맑은 고딕", bold=True, size=10)
    header_fill = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
    )
    cell_font = Font(name="맑은 고딕", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")
    header_align = Alignment(
        wrap_text=True, vertical="center", horizontal="center"
    )

    for filename, tables in file_tables:
        doc_type = detect_document_type(filename)

        # ── 1단계: 유효한 표만 추출하고, 동일 헤더끼리 병합 ──
        merged: dict[str, list[list[str]]] = {}  # header_key → 병합된 행들
        merged_titles: dict[str, str] = {}  # header_key → 대표 제목

        for title, table_lines in tables:
            rows = table_lines_to_rows(table_lines)
            if not rows or is_metadata_table(rows):
                continue

            key = _header_key(rows)
            if key not in merged:
                merged[key] = list(rows)  # 헤더 포함
                merged_titles[key] = title
            else:
                # 헤더(첫 행) 제외하고 데이터만 추가
                merged[key].extend(rows[1:])

        # ── 2단계: 컬럼 재배열 + 시트 생성 ──
        for key, rows in merged.items():
            rows = _reorder_columns(rows, doc_type)
            title = merged_titles[key]

            # 시트 이름 결정
            set_index = _matched_set_index(rows, doc_type)
            if set_index is not None:
                # 본문 데이터 표 — 공공 양식 시트명을 쓴다
                profile = DOCUMENT_PROFILES[doc_type]
                names = profile.get("sheet_names")
                base_name = (
                    names[set_index]
                    if names and set_index < len(names)
                    else profile["sheet_name"]
                )
            elif title:
                # 근거·통계 등 보조 표 — 마크다운 제목을 시트명으로 쓴다
                base_name = title
            else:
                base_name = "Data"

            # 시트 이름 정제 (Excel 제한: 31자, 특수문자 금지)
            sheet_name = re.sub(r'[\\/*?\[\]:]', "", base_name)[:31]
            if not sheet_name:
                sheet_name = "Data"

            # 중복 방지
            original = sheet_name
            dup = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original[:27]}_{dup}"
                dup += 1

            ws = wb.create_sheet(title=sheet_name)

            # 데이터 쓰기
            for row_idx, row in enumerate(rows):
                for col_idx, value in enumerate(row):
                    cell = ws.cell(
                        row=row_idx + 1, column=col_idx + 1, value=value
                    )
                    cell.border = thin_border

                    if row_idx == 0:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                    else:
                        cell.font = cell_font
                        cell.alignment = wrap_align

            # 열 너비 자동 조정
            for col_idx in range(len(rows[0])):
                max_len = 0
                col_letter = get_column_letter(col_idx + 1)
                for row in rows:
                    if col_idx < len(row):
                        length = sum(
                            2 if ord(c) > 127 else 1
                            for c in str(row[col_idx])
                        )
                        max_len = max(max_len, length)
                ws.column_dimensions[col_letter].width = min(
                    max_len + 4, 60
                )

            # 첫 행 고정 (필터용)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

    if not wb.sheetnames:
        ws = wb.create_sheet(title="빈 시트")
        ws.cell(row=1, column=1, value="표 데이터가 없습니다")

    wb.save(output_path)
    return output_path, len(wb.sheetnames)


# ──────────────────────────────────────
# 메인
# ──────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="gx-pm 마크다운 산출물 → xlsx 변환",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python export-xlsx.py ACT-요구사항정의서.md
  python export-xlsx.py --dir ../결과물/
  python export-xlsx.py file1.md file2.md --output merged.xlsx
        """,
    )
    parser.add_argument(
        "files", nargs="*", help="변환할 마크다운 파일 (복수 가능)"
    )
    parser.add_argument(
        "--dir", help="마크다운 파일이 있는 폴더 (폴더 내 모든 .md 처리)"
    )
    parser.add_argument(
        "--output", "-o", help="출력 xlsx 파일 경로 (미지정 시 자동 생성)"
    )
    parser.add_argument(
        "--separate",
        action="store_true",
        help="각 md 파일을 별도 xlsx로 생성 (기본: 하나의 xlsx에 통합)",
    )

    args = parser.parse_args()

    # 입력 파일 수집
    md_files: list[Path] = []

    if args.dir:
        dir_path = Path(args.dir)
        if not dir_path.is_dir():
            print(f"오류: '{args.dir}'은(는) 유효한 폴더가 아닙니다", file=sys.stderr)
            sys.exit(1)
        md_files = sorted(dir_path.glob("*.md"))
    elif args.files:
        for f in args.files:
            p = Path(f)
            if p.is_file():
                md_files.append(p)
            else:
                print(f"경고: '{f}' 파일을 찾을 수 없습니다", file=sys.stderr)
    else:
        # stdin에서 읽기
        print("마크다운 텍스트를 입력하세요 (Ctrl+D로 종료):", file=sys.stderr)
        text = sys.stdin.read()
        tables = parse_markdown_tables(text)
        out = args.output or "output.xlsx"
        result, total_sheets = create_xlsx([("stdin", tables)], out)
        print(f"✅ 생성 완료: {result} ({total_sheets}개 시트)")
        return

    if not md_files:
        print("오류: 변환할 .md 파일이 없습니다", file=sys.stderr)
        sys.exit(1)

    if args.separate:
        # 각 파일을 별도 xlsx로 생성
        for md_file in md_files:
            text = md_file.read_text(encoding="utf-8")
            tables = parse_markdown_tables(text)
            out = md_file.with_suffix(".xlsx")
            if tables:
                result, total_sheets = create_xlsx([(md_file.name, tables)], str(out))
                print(f"✅ {md_file.name} → {result} ({total_sheets}개 시트)")
            else:
                print(f"⚠ {md_file.name}: 표가 없어 건너뜁니다")
    else:
        # 통합 xlsx 생성
        all_tables = []
        for md_file in md_files:
            text = md_file.read_text(encoding="utf-8")
            tables = parse_markdown_tables(text)
            if tables:
                all_tables.append((md_file.name, tables))

        if not all_tables:
            print("오류: 어떤 파일에서도 표를 찾지 못했습니다", file=sys.stderr)
            sys.exit(1)

        if args.output:
            out = args.output
        else:
            # 첫 파일의 시스템코드에서 출력 파일명 생성
            first_name = md_files[0].stem
            prefix_match = re.match(r"^([A-Z]+)-", first_name)
            prefix = prefix_match.group(1) if prefix_match else "output"
            out = f"{prefix}-산출물.xlsx"

        result, total_sheets = create_xlsx(all_tables, out)
        print(f"✅ 생성 완료: {result} ({total_sheets}개 시트)")


if __name__ == "__main__":
    main()
